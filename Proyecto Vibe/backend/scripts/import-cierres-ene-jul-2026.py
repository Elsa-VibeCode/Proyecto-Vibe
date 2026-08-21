#!/usr/bin/env python3
"""Importa cierres Enero–Julio 2026 desde Cierres_BWolven_Enero_Julio_2026_LIMPIO.xlsx.

Por cada mes:
  - Reemplaza egresos del mes
  - Upsert/actualiza facturas (fechaPago, montos, unidad, cliente) desde INGRESOS
"""
from __future__ import annotations

import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import certifi
from openpyxl import load_workbook
from pymongo import MongoClient

EXCEL = Path(
    "/Users/elsaivettedominguezleon/Library/CloudStorage/GoogleDrive-bluewolfadm@gmail.com/"
    "My Drive/GRUPO BLWOLF SAPI DE CV/2026/Reportes GBL/"
    "Cierres_BWolven_Enero_Julio_2026_LIMPIO.xlsx"
)
ENV = Path("/Users/elsaivettedominguezleon/Documents/Repos/Proyecto Vibe/backend/.env")

MESES = {
    "Enero": "2026-01",
    "Febrero": "2026-02",
    "Marzo": "2026-03",
    "Abril": "2026-04",
    "Mayo": "2026-05",
    "Junio": "2026-06",
    "Julio": "2026-07",
}
MESES_ES = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}
UNIDADES = {"Grupo", "Consulting", "Technologies", "Todos"}
ESTADOS = {"ADMINISTRATIVO", "OPERATIVO", "COMERCIAL", "FINANCIERO"}
METODOS = {
    "TRANSFERENCIA",
    "EFECTIVO",
    "TARJETA",
    "SPEI",
    "CHEQUE",
    "DOMICILIACION",
    "OTRO",
}
FACTORES = {
    "IVA_16": 0.16,
    "IVA_8": 0.08,
    "EXENTO": 0.0,
    "CERO": 0.0,
    "ISR_RETENIDO": -0.10467,
}


def round2(n: float) -> float:
    return round(float(n or 0) * 100) / 100


def load_uri() -> str:
    for line in ENV.read_text().splitlines():
        if line.startswith("MONGODB_URI="):
            return line.split("=", 1)[1].strip()
    raise SystemExit("MONGODB_URI no configurada")


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def fecha_utc(v) -> datetime | None:
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day, 12, 0, 0, tzinfo=timezone.utc)
    if isinstance(v, (int, float)) and v > 1000:
        # Excel serial unlikely here with data_only datetimes; ignore
        return None
    texto = str(v or "").strip()
    if not texto:
        return None
    # 5-Ene-2026 / 01-Dic-2025
    m = re.match(r"^(\d{1,2})[-/ ]([A-Za-zÁÉÍÓÚáéíóú\.]+)[-/ ](\d{4})$", texto)
    if m:
        d = int(m.group(1))
        mes_txt = unicodedata.normalize(
            "NFD", m.group(2).lower().replace(".", "")
        )
        mes_txt = "".join(c for c in mes_txt if unicodedata.category(c) != "Mn")[:3]
        y = int(m.group(3))
        mm = MESES_ES.get(mes_txt)
        if mm:
            return datetime(y, mm, d, 12, 0, 0, tzinfo=timezone.utc)
    if re.match(r"^\d{4}-\d{2}-\d{2}", texto):
        y, mth, d = map(int, texto[:10].split("-"))
        return datetime(y, mth, d, 12, 0, 0, tzinfo=timezone.utc)
    return None


def tipo_impuesto(subtotal: float, iva: float) -> str:
    if subtotal <= 0 or abs(iva) < 0.005:
        return "CERO"
    ratio = iva / subtotal
    if abs(ratio - 0.16) < 0.03:
        return "IVA_16"
    if abs(ratio - 0.08) < 0.03:
        return "IVA_8"
    return "CERO"


def enum_val(valor, permitidos, default):
    raw = str(valor or "").strip()
    if raw in permitidos:
        return raw
    up = raw.upper().replace(" ", "_")
    if up in permitidos:
        return up
    for p in permitidos:
        if p.lower() == raw.lower():
            return p
    return default


def normalizar_unidad_factura(valor) -> str | None:
    raw = str(valor or "").strip()
    low = raw.lower()
    if low in {"strategy", "consulting", "consulting / strategy"}:
        return "Consulting"
    if low == "technologies":
        return "Technologies"
    if low in {"grupo", "todos"}:
        return "Grupo"
    return enum_val(raw, UNIDADES - {"Todos"}, None)


def normalizar_unidad_egreso(valor) -> str:
    raw = str(valor or "").strip()
    low = raw.lower()
    if low == "strategy":
        return "Consulting"
    if low == "todos":
        return "Grupo"
    if low == "consulting":
        return "Consulting"
    if low == "technologies":
        return "Technologies"
    if low == "grupo":
        return "Grupo"
    return enum_val(raw, UNIDADES, "Grupo")


def expandir_folios(texto: str, mes: str, cliente: str) -> list[str]:
    partes = [p.strip() for p in re.split(r"\s*/\s*", str(texto or "")) if p.strip()]
    out = []
    for p in partes:
        if p.upper() in {"N/A", "NA", "-"}:
            continue
        if p.upper() == "GAVM":
            slug = re.sub(r"[^A-Za-z0-9]+", "", (cliente or "MARIO")[:12]).upper() or "GAVM"
            out.append(f"GAVM-{mes[5:]}-{slug}")
            continue
        m = re.match(r"^(?:GBL|GB)[-.\s]?(\d+)$", p, re.I)
        if m:
            out.append(f"GBL-{m.group(1)}")
            continue
        m = re.match(r"^(\d+)$", p)
        if m:
            out.append(f"GBL-{m.group(1)}")
            continue
        out.append(p)
    return out


def variantes(folio: str) -> list[str]:
    vals = [folio]
    m = re.match(r"GBL[-.\s]?(\d+)$", folio, re.I)
    if m:
        n = m.group(1)
        vals += [f"GBL-{n}", f"GBL.{n}", f"GB.-{n}", f"GBL {n}"]
    seen, uniq = set(), []
    for v in vals:
        if v not in seen:
            seen.add(v)
            uniq.append(v)
    return uniq


def find_factura(db, folio: str):
    # Preferir activa; si no, reactivar soft-deleted (índice único en noFactura).
    for v in variantes(folio):
        doc = db.facturas.find_one({"noFactura": v, "deletedAt": None})
        if doc:
            return doc
    for v in variantes(folio):
        doc = db.facturas.find_one({"noFactura": v})
        if doc:
            return doc
    return None


def rfc_desde_folio(folio: str, existente=None) -> str:
    if existente and existente.get("rfcEmisor"):
        return existente["rfcEmisor"]
    if str(folio).upper().startswith("GAVM"):
        return "GAVM"
    return "GBL"


def parsear_libro(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    resultado = {}
    for hoja, mes in MESES.items():
        ws = wb[hoja]
        rows = list(ws.iter_rows(min_row=1, max_row=300, max_col=14, values_only=True))
        ingresos, egresos = [], []
        modo = None
        for row in rows:
            c0 = str(row[0] or "").strip()
            if c0 == "INGRESOS":
                modo = "i_title"
                continue
            if c0 == "EGRESOS":
                modo = "e_title"
                continue
            if modo == "i_title" and "Fecha de facturación" in c0:
                modo = "ingresos"
                continue
            if modo == "e_title" and c0.startswith("Núm"):
                modo = "egresos"
                continue

            if modo == "ingresos":
                folio_raw = str(row[1] or "").strip()
                if not folio_raw or folio_raw.upper().startswith("TOTAL"):
                    continue
                fecha_pago = fecha_utc(row[2])
                if not fecha_pago:
                    continue
                subtotal = round2(num(row[7]))
                iva = round2(num(row[8]))
                total = round2(num(row[9]))
                if total <= 0 and subtotal <= 0:
                    continue
                if total < 1 and folio_raw.upper() in {"N/A", "NA"}:
                    continue  # compensaciones bancarias mínimas
                if total <= 0:
                    total = round2(subtotal + iva)
                cliente = str(row[5] or "").strip()
                folios = expandir_folios(folio_raw, mes, cliente)
                if not folios:
                    # N/A con monto real → folio sintético
                    folios = [f"CIERRE-{mes}-{fecha_pago.strftime('%d')}-{round(total*100)}"]
                fecha_fact = fecha_utc(row[0]) or fecha_pago
                estatus = str(row[11] or "PAGADO").strip().upper() or "PAGADO"
                if "PEND" in estatus:
                    estatus = "PENDIENTE"
                elif "PARCIAL" in estatus:
                    estatus = "PARCIAL"
                else:
                    estatus = "PAGADO"
                ingresos.append(
                    {
                        "foliosRaw": folio_raw,
                        "folios": folios,
                        "fechaFacturacion": fecha_fact,
                        "fechaPago": fecha_pago,
                        "unidad": normalizar_unidad_factura(row[3]),
                        "cliente": cliente,
                        "concepto": str(row[6] or "").strip(),
                        "subtotal": subtotal,
                        "iva": iva,
                        "total": total,
                        "estatusPago": estatus,
                        "mes": mes,
                    }
                )
            elif modo == "egresos":
                fecha = fecha_utc(row[1])
                if not fecha:
                    continue
                proveedor = str(row[9] or "").strip()
                concepto = str(row[10] or "").strip()
                subtotal = round2(num(row[11]))
                iva = round2(num(row[12]))
                if not proveedor and subtotal <= 0:
                    continue
                if not proveedor:
                    continue
                no_fact = str(row[8] or "").strip()
                if no_fact.upper() in {"N/A", "NA"}:
                    no_fact = ""
                tipo = tipo_impuesto(subtotal, iva)
                impuesto = round2(subtotal * FACTORES[tipo])
                # Prefer Excel IVA when present; keep total = subtotal + iva Excel
                if abs(iva) > 0.001:
                    impuesto = iva
                total = round2(subtotal + impuesto)
                metodo = enum_val(str(row[5] or "").upper(), METODOS, "TRANSFERENCIA")
                egresos.append(
                    {
                        "fechaGasto": fecha,
                        "proyecto": str(row[6] or "").strip(),
                        "estadoResultado": enum_val(row[7], ESTADOS, "ADMINISTRATIVO"),
                        "unidad": normalizar_unidad_egreso(row[2]),
                        "tipoGasto": str(row[3] or "OTROS").strip().upper() or "OTROS",
                        "tipoSubgasto": str(row[4] or "").strip(),
                        "metodoPago": metodo,
                        "noFactura": no_fact,
                        "proveedor": proveedor,
                        "concepto": concepto or proveedor,
                        "subtotal": subtotal,
                        "tipoImpuesto": tipo,
                        "impuesto": impuesto,
                        "total": total,
                        "mes": mes,
                        "esTransferLatam": "LATAM" in (concepto or "").upper(),
                        "createdAt": datetime.now(timezone.utc),
                        "updatedAt": datetime.now(timezone.utc),
                    }
                )
        resultado[mes] = {"hoja": hoja, "ingresos": ingresos, "egresos": egresos}
    wb.close()
    return resultado


def importar_mes(db, mes: str, ingresos, egresos):
    print(f"\n=== {mes} ({len(ingresos)} ingresos, {len(egresos)} egresos) ===")
    tot_i = round2(sum(i["total"] for i in ingresos))
    tot_e = round2(sum(e["total"] for e in egresos))
    print(f"  Excel: ingresos ${tot_i:,.2f} · egresos ${tot_e:,.2f}")

    borrados = db.egresos.delete_many({"mes": mes})
    if egresos:
        db.egresos.insert_many(egresos)
    print(f"  Egresos: borrados {borrados.deleted_count}, insertados {len(egresos)}")
    for e in db.egresos.aggregate(
        [
            {"$match": {"mes": mes}},
            {"$group": {"_id": "$unidad", "n": {"$sum": 1}, "total": {"$sum": "$total"}}},
            {"$sort": {"_id": 1}},
        ]
    ):
        print(f"    {e['_id']}: {e['n']}  ${round2(e['total']):,.2f}")

    actualizadas = creadas = montos = 0
    for row in ingresos:
        folios = row["folios"]
        aplicar_monto = len(folios) == 1
        for folio in folios:
            doc = find_factura(db, folio)
            payload = {
                "fechaPago": row["fechaPago"] if row["estatusPago"] == "PAGADO" else (doc or {}).get("fechaPago"),
                "estatusPago": row["estatusPago"],
                "estatusEnvio": "ENVIADA",
                "cliente": row["cliente"] or (doc.get("cliente") if doc else folio),
                "concepto": row["concepto"] or (doc.get("concepto") if doc else ""),
                "unidad": row["unidad"] or (doc.get("unidad") if doc else "Grupo"),
                "unidadManual": True,
                "rfcEmisor": rfc_desde_folio(folio, doc),
                "updatedAt": datetime.now(timezone.utc),
            }
            if row["estatusPago"] != "PAGADO":
                payload["fechaPago"] = None
            if aplicar_monto:
                payload.update(
                    {
                        "subtotal": row["subtotal"],
                        "iva": row["iva"],
                        "total": row["total"],
                    }
                )
            if doc:
                prev = round2(doc.get("total") or 0)
                # No pisar fechaFacturacion si ya existe; reactivar si estaba borrada
                update = {"$set": payload, "$unset": {"deletedAt": ""}}
                db.facturas.update_one({"_id": doc["_id"]}, update)
                actualizadas += 1
                if aplicar_monto and abs(prev - row["total"]) > 0.05:
                    montos += 1
            else:
                payload.update(
                    {
                        "noFactura": folio,
                        "fechaFacturacion": row["fechaFacturacion"],
                        "mes": row["mes"],
                        "origen": "excel-migracion",
                        "metodoPago": "PUE",
                        "clasificacionAuto": False,
                        "subtotal": row["subtotal"] if aplicar_monto else 0,
                        "iva": row["iva"] if aplicar_monto else 0,
                        "total": row["total"] if aplicar_monto else 0,
                        "createdAt": datetime.now(timezone.utc),
                        "deletedAt": None,
                    }
                )
                if row["estatusPago"] != "PAGADO":
                    payload["fechaPago"] = None
                try:
                    db.facturas.insert_one(payload)
                    creadas += 1
                except Exception as exc:
                    # Carrera / índice único: reintentar como update
                    if "E11000" not in str(exc):
                        raise
                    existente = find_factura(db, folio)
                    if not existente:
                        raise
                    db.facturas.update_one(
                        {"_id": existente["_id"]},
                        {"$set": payload, "$unset": {"deletedAt": ""}},
                    )
                    actualizadas += 1
    print(f"  Ingresos: {actualizadas} actualizadas, {creadas} creadas, {montos} montos alineados")
    return tot_i, tot_e


def main():
    data = parsear_libro(EXCEL)
    db = MongoClient(
        load_uri(), tlsCAFile=certifi.where(), serverSelectionTimeoutMS=30000
    ).get_default_database()

    suma_i = suma_e = 0
    for mes in sorted(data.keys()):
        ti, te = importar_mes(db, mes, data[mes]["ingresos"], data[mes]["egresos"])
        suma_i += ti
        suma_e += te

    print("\n=== RESUMEN ENE–JUL ===")
    print(f"Ingresos Excel (con IVA): ${round2(suma_i):,.2f}")
    print(f"Egresos Excel (con IVA):  ${round2(suma_e):,.2f}")

    for mes in sorted(MESES.values()):
        eg = list(
            db.egresos.aggregate(
                [
                    {"$match": {"mes": mes}},
                    {"$group": {"_id": None, "n": {"$sum": 1}, "t": {"$sum": "$total"}}},
                ]
            )
        )
        y, m = map(int, mes.split("-"))
        ini = datetime(y, m, 1)
        fin = datetime(y + (m == 12), 1 if m == 12 else m + 1, 1)
        ing = list(
            db.facturas.aggregate(
                [
                    {
                        "$match": {
                            "deletedAt": None,
                            "fechaPago": {"$gte": ini, "$lt": fin},
                            "estatusPago": "PAGADO",
                        }
                    },
                    {"$group": {"_id": None, "n": {"$sum": 1}, "t": {"$sum": "$total"}}},
                ]
            )
        )
        print(
            f"  {mes}: egresos {eg[0]['n'] if eg else 0} ${round2(eg[0]['t'] if eg else 0):,.2f}"
            f" · cobrado {ing[0]['n'] if ing else 0} ${round2(ing[0]['t'] if ing else 0):,.2f}"
        )


if __name__ == "__main__":
    main()
