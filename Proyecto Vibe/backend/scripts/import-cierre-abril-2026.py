#!/usr/bin/env python3
"""Importa el cierre Abril 2026 (egresos + fechas de cobro) desde el Excel de GBL.

Reemplaza egresos mes=2026-04 y actualiza fechaPago/estatusPago de las facturas
listadas en INGRESOS ABRIL. No recrea CFDIs: si el folio no existe, lo reporta.
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

import certifi
from openpyxl import load_workbook
from pymongo import MongoClient

EXCEL = Path(
    "/Users/elsaivettedominguezleon/Library/CloudStorage/GoogleDrive-bluewolfadm@gmail.com/"
    "My Drive/GRUPO BLWOLF SAPI DE CV/2026/Reportes GBL/ABRIL/"
    "Cierre Abril2026 Ingresos-Egresos.xlsx"
)
ENV = Path("/Users/elsaivettedominguezleon/Documents/Repos/Proyecto Vibe/backend/.env")
MES = "2026-04"
UNIDADES = {"Grupo", "Consulting", "Technologies", "Todos"}
ESTADOS = {"ADMINISTRATIVO", "OPERATIVO", "COMERCIAL", "FINANCIERO"}
METODOS = {
    "TRANSFERENCIA",
    "EFECTIVO",
    "TARJETA",
    "SPEI",
    "CHEQUE",
    "DOMICILIACION",
    "OTRO",
}
FACTORES = {
    "IVA_16": 0.16,
    "IVA_8": 0.08,
    "EXENTO": 0.0,
    "CERO": 0.0,
    "ISR_RETENIDO": -0.10467,
}


def round2(n: float) -> float:
    return round(float(n or 0) * 100) / 100


def load_uri() -> str:
    for line in ENV.read_text().splitlines():
        if line.startswith("MONGODB_URI="):
            return line.split("=", 1)[1].strip()
    raise SystemExit("MONGODB_URI no configurada")


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def fecha_utc(v) -> datetime | None:
    if isinstance(v, datetime):
        return datetime(v.year, v.month, v.day, 12, 0, 0, tzinfo=timezone.utc)
    return None


def tipo_impuesto(subtotal: float, iva: float) -> str:
    if subtotal <= 0 or iva == 0:
        return "CERO"
    ratio = iva / subtotal
    if abs(ratio - 0.16) < 0.03:
        return "IVA_16"
    if abs(ratio - 0.08) < 0.03:
        return "IVA_8"
    return "CERO"


def enum_val(valor, permitidos, default):
    raw = str(valor or "").strip()
    if raw in permitidos:
        return raw
    up = raw.upper().replace(" ", "_")
    if up in permitidos:
        return up
    for p in permitidos:
        if p.lower() == raw.lower():
            return p
    return default


def parse_sheet(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=420, max_col=15, values_only=True))
    wb.close()

    egresos = []
    ingresos = []
    seccion = None
    for i, row in enumerate(rows, 1):
        c1 = str(row[0] or "").strip()
        if c1 == "EGRESOS ABRIL":
            seccion = "egresos_title"
            continue
        if c1 == "Num" and str(row[1] or "").startswith("Fecha"):
            seccion = "egresos"
            continue
        if c1 == "INGRESOS ABRIL":
            seccion = "ingresos_title"
            continue
        if c1.startswith("No. Factura"):
            seccion = "ingresos"
            continue
        if seccion == "egresos":
            if not row[1]:
                continue
            fecha = fecha_utc(row[1])
            if not fecha:
                continue
            proveedor = str(row[10] or "").strip()
            concepto = str(row[11] or "").strip()
            subtotal = round2(num(row[12]))
            iva = round2(num(row[13]))
            if not proveedor or subtotal <= 0:
                continue
            no_fact = str(row[9] or "").strip()
            if no_fact.upper() in {"N/A", "NA"}:
                no_fact = ""
            tipo = tipo_impuesto(subtotal, iva)
            factor = FACTORES[tipo]
            impuesto = round2(subtotal * factor)
            total = round2(subtotal + impuesto)
            egresos.append(
                {
                    "fechaGasto": fecha,
                    "proyecto": str(row[7] or "").strip(),
                    "estadoResultado": enum_val(row[8], ESTADOS, "ADMINISTRATIVO"),
                    "unidad": enum_val(row[2], UNIDADES, "Grupo"),
                    "tipoGasto": str(row[3] or "OTROS").strip().upper() or "OTROS",
                    "tipoSubgasto": str(row[4] or "").strip(),
                    "metodoPago": enum_val(str(row[5] or "").upper(), METODOS, "TRANSFERENCIA"),
                    "noFactura": no_fact,
                    "proveedor": proveedor,
                    "concepto": concepto or proveedor,
                    "subtotal": subtotal,
                    "tipoImpuesto": tipo,
                    "impuesto": impuesto,
                    "total": total,
                    "mes": MES,
                    "esTransferLatam": "LATAM" in (concepto or "").upper(),
                    "createdAt": datetime.now(timezone.utc),
                    "updatedAt": datetime.now(timezone.utc),
                }
            )
        elif seccion == "ingresos":
            folio = str(row[0] or "").strip()
            if not folio or folio.startswith("TOTAL"):
                continue
            fecha = fecha_utc(row[1])
            if not fecha:
                continue
            ingresos.append(
                {
                    "foliosRaw": folio,
                    "folios": expandir_folios(folio),
                    "fechaPago": fecha,
                    "unidad": enum_val(row[2], UNIDADES, None),
                    "cliente": str(row[4] or "").strip(),
                    "concepto": str(row[7] or "").strip(),
                    "subtotal": round2(num(row[12])),
                    "iva": round2(num(row[13])),
                    "total": round2(num(row[14])),
                }
            )
    return egresos, ingresos


def expandir_folios(texto: str) -> list[str]:
    partes = [p.strip() for p in re.split(r"\s*/\s*", texto) if p.strip()]
    out = []
    for p in partes:
        m = re.match(r"^(?:GBL[-.\s]?)?(\d+)$", p, re.I)
        if m:
            out.append(f"GBL-{m.group(1)}")
            continue
        out.append(p)
    return out


def variantes(folio: str) -> list[str]:
    vals = [folio]
    m = re.match(r"GBL[-.\s]?(\d+)$", folio, re.I)
    if m:
        n = m.group(1)
        vals += [f"GBL-{n}", f"GBL.{n}", f"GB.-{n}", f"GBL {n}"]
    # unique preserve order
    seen = set()
    uniq = []
    for v in vals:
        if v not in seen:
            seen.add(v)
            uniq.append(v)
    return uniq


def find_factura(db, folio: str):
    for v in variantes(folio):
        doc = db.facturas.find_one({"noFactura": v, "deletedAt": None})
        if doc:
            return doc
    return None


def main():
    egresos, ingresos = parse_sheet(EXCEL)
    print(f"Excel: {len(egresos)} egresos, {len(ingresos)} ingresos")
    tot_e = round2(sum(e["total"] for e in egresos))
    tot_i = round2(sum(i["total"] for i in ingresos))
    print(f"  totales Excel  egresos ${tot_e:,.2f}  ingresos ${tot_i:,.2f}")

    db = MongoClient(
        load_uri(), tlsCAFile=certifi.where(), serverSelectionTimeoutMS=25000
    ).get_default_database()

    antes_e = db.egresos.count_documents({"mes": MES})
    borrados = db.egresos.delete_many({"mes": MES})
    if egresos:
        db.egresos.insert_many(egresos)
    print(f"Egresos {MES}: borrados {borrados.deleted_count} (había {antes_e}), insertados {len(egresos)}")

    por_u = {}
    for e in db.egresos.aggregate(
        [
            {"$match": {"mes": MES}},
            {"$group": {"_id": "$unidad", "n": {"$sum": 1}, "total": {"$sum": "$total"}}},
            {"$sort": {"_id": 1}},
        ]
    ):
        por_u[e["_id"]] = (e["n"], round2(e["total"]))
        print(f"  {e['_id']}: {e['n']}  ${round2(e['total']):,.2f}")

    creadas = 0
    actualizadas = 0
    montos = 0
    for row in ingresos:
        folios = row["folios"]
        # Una fila del cierre = un cobro. Si trae un solo folio, el monto del Excel
        # es la fuente. Si trae varios, solo se marca pagada (el CFDI ya trae el total).
        aplicar_monto = len(folios) == 1
        for folio in folios:
            doc = find_factura(db, folio)
            payload = {
                "fechaPago": row["fechaPago"],
                "estatusPago": "PAGADO",
                "estatusEnvio": "ENVIADA",
                "cliente": row["cliente"] or (doc.get("cliente") if doc else folio),
                "concepto": row["concepto"] or (doc.get("concepto") if doc else ""),
                "unidad": row["unidad"] or (doc.get("unidad") if doc else "Grupo"),
                "unidadManual": True,
                "rfcEmisor": (doc.get("rfcEmisor") if doc else "GBL") or "GBL",
                "updatedAt": datetime.now(timezone.utc),
            }
            if aplicar_monto:
                payload.update(
                    {
                        "subtotal": row["subtotal"],
                        "iva": row["iva"],
                        "total": row["total"],
                    }
                )
            if doc:
                prev = round2(doc.get("total") or 0)
                db.facturas.update_one({"_id": doc["_id"]}, {"$set": payload})
                actualizadas += 1
                if aplicar_monto and abs(prev - row["total"]) > 0.05:
                    montos += 1
                    print(f"  monto {folio}: ${prev:,.2f} → ${row['total']:,.2f}")
            else:
                payload.update(
                    {
                        "noFactura": folio,
                        "fechaFacturacion": row["fechaPago"],
                        "mes": MES,
                        "origen": "excel-migracion",
                        "metodoPago": "PUE",
                        "clasificacionAuto": False,
                        "subtotal": row["subtotal"] if aplicar_monto else 0,
                        "iva": row["iva"] if aplicar_monto else 0,
                        "total": row["total"] if aplicar_monto else 0,
                        "createdAt": datetime.now(timezone.utc),
                        "deletedAt": None,
                    }
                )
                db.facturas.insert_one(payload)
                creadas += 1
                print(f"  creada {folio} ${row['total']:,.2f} ({row['unidad']})")

    print(f"Ingresos: {actualizadas} actualizadas, {creadas} creadas, {montos} montos alineados al cierre")

    ini = datetime(2026, 4, 1)
    fin = datetime(2026, 5, 1)
    cobrado = list(
        db.facturas.aggregate(
            [
                {
                    "$match": {
                        "deletedAt": None,
                        "fechaPago": {"$gte": ini, "$lt": fin},
                        "rfcEmisor": {"$ne": "GAVM"},
                    }
                },
                {"$group": {"_id": "$unidad", "n": {"$sum": 1}, "total": {"$sum": "$total"}}},
                {"$sort": {"_id": 1}},
            ]
        )
    )
    print("Cobrado GBL en sistema (sin Mario/GAVM):")
    suma = 0
    for r in cobrado:
        print(f"  {r['_id']}: {r['n']}  ${round2(r['total']):,.2f}")
        suma += r["total"]
    print(f"  total ${round2(suma):,.2f}  (Excel ingresos ${tot_i:,.2f})")


if __name__ == "__main__":
    main()
